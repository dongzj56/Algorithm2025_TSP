import openpyxl
from pathlib import Path

def load_best_solutions(txt_path):
    best_dict = {}
    with open(txt_path, 'r', encoding='utf-8') as f:
        for line in f:
            if ':' in line:
                name, val = line.strip().split(':')
                best_dict[name.strip()] = int(val.strip())
    return best_dict

def update_excel_with_best(xlsx_path, best_dict):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # 获取列索引
    header = [cell.value for cell in ws[1]]
    file_col = header.index('file') + 1
    if 'best' in header:
        best_col = header.index('best') + 1
    else:
        best_col = len(header) + 1
        ws.cell(row=1, column=best_col, value='best')

    for row in range(2, ws.max_row + 1):
        filename = str(ws.cell(row=row, column=file_col).value).strip()
        key = Path(filename).stem  # 去掉扩展名
        if key in best_dict:
            ws.cell(row=row, column=best_col, value=best_dict[key])

    wb.save(xlsx_path)
    print(f"✅ 最优路径已写入 Excel：{xlsx_path}")

if __name__ == "__main__":
    txt_path = rf"C:\Users\dongz\Desktop\Algorithm_TSP_2025\data\tsp\bestSolutions.txt"
    xlsx_path = rf"C:\Users\dongz\Desktop\Algorithm_TSP_2025\output\tsp_sizes.xlsx"
    
    best_dict = load_best_solutions(txt_path)
    update_excel_with_best(xlsx_path, best_dict)
