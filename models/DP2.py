import math, time
from pathlib import Path
import numpy as np
import openpyxl

# ---------------- DP 类（保持不变） ----------------
class DP(object):
    def __init__(self, num_city, num_total, iteration, data):
        self.num_city = num_city
        self.location = data
        self.dis_mat = self.compute_dis_mat(num_city, data)

    def compute_dis_mat(self, num_city, location):
        dis_mat = np.zeros((num_city, num_city))
        for i in range(num_city):
            for j in range(num_city):
                if i == j:
                    dis_mat[i][j] = np.inf
                    continue
                a, b = location[i], location[j]
                dis_mat[i][j] = np.sqrt(np.sum((a - b) ** 2))
        return dis_mat

    def compute_pathlen(self, path):
        dis = 0.0
        for i in range(len(path) - 1):
            dis += self.dis_mat[path[i]][path[i + 1]]
        dis += self.dis_mat[path[-1]][path[0]]  # 回到起点
        return dis

    def run(self):
        rest = list(range(1, self.num_city))
        path = [0]
        while rest:
            c = rest.pop(0)
            if len(path) == 1:
                path.append(c)
                continue
            best_idx, best_inc = 0, math.inf
            for i in range(len(path)):
                a = path[i - 1]
                b = path[i]
                inc = self.dis_mat[a][c] + self.dis_mat[c][b] - self.dis_mat[a][b]
                if inc < best_inc:
                    best_inc, best_idx = inc, i
            path.insert(best_idx, c)
        best_len = self.compute_pathlen(path)
        return path, best_len

# ---------------- 读取 TSPLIB ----------------

def read_tsp(path):
    """仅解析含 NODE_COORD_SECTION 的坐标型 TSP；遇显式矩阵直接抛错。"""
    lines = Path(path).read_text().splitlines()
    idx = next((i for i, l in enumerate(lines) if l.strip().upper() == "NODE_COORD_SECTION"), None)
    if idx is None:
        raise ValueError(f"{path} 缺少 NODE_COORD_SECTION")
    coords = []
    for line in lines[idx + 1:]:
        line = line.strip()
        if not line or line.upper() == "EOF":
            break
        parts = [float(p) for p in line.split() if p]
        coords.append(parts[1:])  # 丢弃序号
    return np.array(coords)

# ---------------- 按 Excel 表批量处理 ----------------

def solve_from_excel(xlsx_path, tsp_dir):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # 检查/添加表头
    headers = [cell.value for cell in ws[1]]
    if 'dp_route' not in headers:
        ws.cell(row=1, column=len(headers) + 1, value='dp_route')
        ws.cell(row=1, column=len(headers) + 2, value='dp_time')

    route_col = headers.index('dp_route') + 1 if 'dp_route' in headers else len(headers) + 1
    time_col = headers.index('dp_time') + 1 if 'dp_time' in headers else len(headers) + 2

    for row in range(2, ws.max_row + 1):
        tsp_name = str(ws.cell(row=row, column=1).value).strip()
        tsp_path = Path(tsp_dir) / tsp_name

        if not tsp_path.exists():
            ws.cell(row=row, column=route_col, value="File not found")
            ws.cell(row=row, column=time_col, value=None)
            continue

        # 记录单个实例开始时间（包括读取、建模、求解全过程）
        inst_start = time.perf_counter()

        try:
            coords = read_tsp(tsp_path)
        except ValueError as e:
            print(f"[跳过] {tsp_name}: {e}")
            ws.cell(row=row, column=route_col, value="No coords section")
            ws.cell(row=row, column=time_col, value=None)
            continue

        if coords.size == 0:
            ws.cell(row=row, column=route_col, value="Empty coords")
            ws.cell(row=row, column=time_col, value=None)
            continue

        model = DP(num_city=len(coords), num_total=0, iteration=0, data=coords)
        route, path_len = model.run()

        inst_end = time.perf_counter()
        total_time = round(inst_end - inst_start, 4)  # 总用时（秒）

        ws.cell(row=row, column=route_col, value=" ".join(map(str, route)))
        ws.cell(row=row, column=time_col, value=total_time)
        print(f"{tsp_name:<20} | length={path_len:.2f} | time={total_time} s")

    wb.save(xlsx_path)
    print(f"已更新：{xlsx_path}")

# ---------------- 调用 ----------------
if __name__ == "__main__":
    solve_from_excel(
        xlsx_path=rf"C:\Users\dongz\Desktop\Algorithm_TSP_2025\output\tsp_sizes.xlsx",  # Excel 路径
        tsp_dir=rf"C:\Users\dongz\Desktop\Algorithm_TSP_2025\data\tsp"                 # .tsp 文件所在目录
    )
